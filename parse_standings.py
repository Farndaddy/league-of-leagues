#!/usr/bin/env python3
"""Parse LOL_Standings_Champions_Earnings.xlsx into structured JSON."""
import openpyxl, json, sys

XLSX_PATH = '/sessions/fervent-serene-maxwell/mnt/uploads/LOL_Standings_Champions_Earnings.xlsx'

# Normalize owner names — fix known aliases in the spreadsheet
OWNER_ALIASES = {
    'andrew & john': 'Andrew & Timpe',
    'larry & mohit': 'Larry & Mo',
    'moses':         'Moses Joseph',
    'crump - champ': 'Chris Crump',
}

def norm_owner(name):
    if not name: return name
    lo = str(name).strip().lower()
    for alias, canonical in OWNER_ALIASES.items():
        if alias in lo:
            return canonical
    return str(name).strip()

# Map team name variants to canonical names
TEAM_ALIASES = {
    'fcc':               'Thanos',
    'fcc/thanos':        'Thanos',
    'the champ':         'CHAMP',
    'lair of maverick':  'Lair of the Maverick',
    'moviepoopshoot.com':'MoviePoopShoot',
    'turbo team':        'Turbo Team',
}
def norm_team(name):
    if not name: return name
    lo = str(name).strip().lower()
    for alias, canonical in TEAM_ALIASES.items():
        if lo == alias:
            return canonical
    return str(name).strip()

def parse_season_sheet(ws, year_key):
    """Parse a season sheet — standings rows + earnings rows."""
    rows = list(ws.iter_rows(values_only=True))

    standings = []
    earnings  = []
    section   = 'standings'
    sport_section = None

    for row in rows:
        r = [v for v in row[:7]]
        # Pad to 7
        while len(r) < 7: r.append(None)
        rnk, label, c2, c3, c4, c5, c6 = r

        if label is None and c2 is None:
            continue

        label_str = str(label).strip() if label else ''
        c2_str    = str(c2).strip()    if c2    else ''

        # Detect section headers
        if 'earnings' in label_str.lower() and rnk is None:
            section = 'earnings'
            sport_section = None
            if 'nfl' in label_str.lower():  sport_section = 'NFL'
            elif 'nba' in label_str.lower(): sport_section = 'NBA'
            elif 'mlb' in label_str.lower(): sport_section = 'MLB'
            continue
        if section == 'earnings' and label_str.lower() in ('nfl earnings','nba earnings','mlb earnings'):
            if 'nfl' in label_str.lower():  sport_section = 'NFL'
            elif 'nba' in label_str.lower(): sport_section = 'NBA'
            elif 'mlb' in label_str.lower(): sport_section = 'MLB'
            continue

        # Skip header rows
        if isinstance(rnk, str) and 'rnk' in rnk.lower():
            continue
        if label_str.lower() in ('winner', 'sub total', 'amount'):
            continue
        if label_str.startswith('Note'):
            continue
        if label_str.lower() in ('historical standings', '2021-2024 seasons',
                                  'overall points', 'nfl points', 'nba points', 'mlb points'):
            continue

        if section == 'standings':
            if isinstance(rnk, (int, float)) and label:
                tiebreaker = str(c6).strip() if c6 else None
                standings.append({
                    'rank':       int(rnk),
                    'team':       norm_team(label_str),
                    'nfl_pts':    int(c2) if isinstance(c2, (int,float)) else None,
                    'nba_pts':    int(c3) if isinstance(c3, (int,float)) else None,
                    'mlb_pts':    int(c4) if isinstance(c4, (int,float)) else None,
                    'total_pts':  int(c5) if isinstance(c5, (int,float)) else None,
                    'tiebreaker': tiebreaker if tiebreaker and 'tie' in tiebreaker.lower() else None,
                })

        elif section == 'earnings':
            if label_str and isinstance(c2, (int,float)):
                # c3 is Winner string
                winner_raw = str(c3).strip() if c3 else ''
                earnings.append({
                    'place':       label_str,   # e.g. "1st Place", "2nd Place", "Reg Season"
                    'amount':      int(c2),
                    'winner':      winner_raw,
                    'sport':       sport_section,  # None = overall
                })

    return {'standings': standings, 'earnings': earnings}

def parse_cumulative(ws):
    """Parse the 2021-2024 cumulative tab."""
    rows = list(ws.iter_rows(values_only=True))
    result = []
    for row in rows:
        rnk, label, c2, c3, c4, c5 = (list(row[:6]) + [None]*6)[:6]
        if isinstance(rnk, (int,float)) and label:
            result.append({
                'rank':      int(rnk),
                'team':      norm_team(str(label).strip()),
                'nfl_pts':   int(c2) if isinstance(c2,(int,float)) else None,
                'nba_pts':   int(c3) if isinstance(c3,(int,float)) else None,
                'mlb_pts':   int(c4) if isinstance(c4,(int,float)) else None,
                'total_pts': int(c5) if isinstance(c5,(int,float)) else None,
            })
    return result

def parse_champions(ws):
    """Parse the Champions tab."""
    rows = list(ws.iter_rows(values_only=True))
    section = None
    result  = {'overall': [], 'nfl': [], 'nba': [], 'mlb': []}

    for row in rows:
        v0, v1, v2 = (list(row[:3]) + [None,None,None])[:3]
        label = str(v0).strip() if v0 else ''

        if label.lower() == 'champions':       section = 'overall'; continue
        if label.lower() == 'nfl champs':      section = 'nfl';     continue
        if label.lower() == 'nba champs':      section = 'nba';     continue
        if label.lower() == 'mlb champs':      section = 'mlb';     continue
        if label.lower() in ('overall points',): continue

        if isinstance(v0, (int,float)) and section:
            year  = int(v0)
            owner = norm_owner(v1) if v1 else None
            pts   = int(v2) if isinstance(v2,(int,float)) else None
            entry = {'year': year, 'owner': owner}
            if pts is not None: entry['points'] = pts
            result[section].append(entry)

    return result

def parse_career_earnings(ws):
    """Parse Career Earnings tab — top summary table only."""
    rows = list(ws.iter_rows(values_only=True))
    earners = []
    for row in rows:
        v = list(row[:6])
        owner, total, y2022, y2023, y2024, y2025 = (v + [None]*6)[:6]
        if not owner or not isinstance(owner, str): continue
        if owner.strip().lower() in ('owner', 'year', 'earnings', 'team'): continue
        if isinstance(total, (int,float)):
            earners.append({
                'owner':   norm_owner(owner),
                'total':   int(total),
                'by_year': {
                    '2022': int(y2022) if isinstance(y2022,(int,float)) else 0,
                    '2023': int(y2023) if isinstance(y2023,(int,float)) else 0,
                    '2024': int(y2024) if isinstance(y2024,(int,float)) else 0,
                    '2025': int(y2025) if isinstance(y2025,(int,float)) else 0,
                }
            })
    # Sort by total desc
    earners.sort(key=lambda x: x['total'], reverse=True)
    return earners

def run():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # Season sheets: key = end year
    seasons = {}
    season_map = {
        '2021-2022 Standings': 2022,
        '2022-2023 Season':    2023,
        '2023-2024 Season':    2024,
        '2024-25 Season':      2025,
    }
    for sheet_name, year_key in season_map.items():
        if sheet_name in wb.sheetnames:
            seasons[year_key] = parse_season_sheet(wb[sheet_name], year_key)
            print(f"  {sheet_name}: {len(seasons[year_key]['standings'])} teams, "
                  f"{len(seasons[year_key]['earnings'])} earnings rows", file=sys.stderr)

    cumulative = parse_cumulative(wb['2021-2024 Seasons'])
    print(f"  Cumulative: {len(cumulative)} teams", file=sys.stderr)

    champions = parse_champions(wb['Champions'])
    print(f"  Champions: overall={len(champions['overall'])}, "
          f"nfl={len(champions['nfl'])}, nba={len(champions['nba'])}, "
          f"mlb={len(champions['mlb'])}", file=sys.stderr)

    career = parse_career_earnings(wb['Career Earnings'])
    print(f"  Career earnings: {len(career)} owners", file=sys.stderr)

    return {
        'seasons':         seasons,
        'cumulative':      cumulative,
        'champions':       champions,
        'careerEarnings':  career,
    }

if __name__ == '__main__':
    d = run()
    print(json.dumps(d, indent=2))
