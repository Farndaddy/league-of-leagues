#!/usr/bin/env python3
"""Parse LOL Draft Boards.xlsx into structured JSON."""
import openpyxl, json, re, sys

XLSX_PATH = '/sessions/fervent-serene-maxwell/mnt/uploads/LOL Draft Boards-66f77893.xlsx'

KEEPER_ROUNDS = set(range(33, 42))   # rounds 33-41 = keepers
KEEPER_YEARS  = {2022, 2023, 2024, 2025}

SPORT_COLORS = {
    'EAD1DC':'NFL','E6B8C8':'NFL','D5A6BD':'NFL','EA9999':'NFL','F4CCCC':'NFL','FFB6C1':'NFL',
    'C9DAF8':'NBA','A4C2F4':'NBA','9FC5E8':'NBA','6D9EEB':'NBA','CFE2F3':'NBA','BDD7EE':'NBA',
    'ADD8E6':'NBA','CEE1F2':'NBA','C0D7F0':'NBA','B4C7E7':'NBA','DAE8F5':'NBA',
    'FFF2CC':'MLB','FFE599':'MLB','FFD966':'MLB','FFFF00':'MLB','FFFACD':'MLB','FFF3C1':'MLB',
    'FCEBD7':'MLB','FCE5CD':'MLB','FFF9E6':'MLB',
    '999999':'SUPP','B7B7B7':'SUPP','CCCCCC':'SUPP','D9D9D9':'SUPP','B6D7A8':'SUPP',
    'C9C9C9':'SUPP','A9A9A9':'SUPP','EFEFEF':'SUPP','808080':'SUPP','7F7F7F':'SUPP',
    '93C47D':'SUPP','6AA84F':'SUPP',
}

def keeper_sport(i):
    return 'NFL' if i < 3 else ('NBA' if i < 6 else 'MLB')

def get_hex(cell):
    try:
        fg = cell.fill.fgColor
        if fg.type == 'rgb' and fg.rgb not in ('00000000','FF000000','FFFFFFFF',''):
            return fg.rgb[2:].upper()
    except: pass
    return None

def sport_of(h): return SPORT_COLORS.get((h or '').upper())
def is_bold(cell):
    try: return bool(cell.font and cell.font.bold)
    except: return False
def cs(v): return str(v).strip() if v is not None else ''
def extract_init(s):
    m = re.search(r'\(([^)]+)\)', s)
    return m.group(1).strip() if m else None
def clean_owner(s): return re.sub(r'\s*\([^)]+\)', '', s).strip()

def parse_cell(cell, val_str, is_future=False):
    hex_c = get_hex(cell)
    sport = sport_of(hex_c)
    bold  = is_bold(cell)

    if is_future and hex_c and hex_c.upper() == 'FF9900':
        return {'player':None, 'sport':None, 'color':hex_c,
                'owned_by':val_str.rstrip('*').strip(), 'is_traded':True}

    if val_str.upper().startswith('KPR '):
        return {'player':val_str[4:].strip(), 'sport':sport, 'color':hex_c, 'is_keeper':True}

    if bold and ' - ' in val_str:
        parts = val_str.split(' - ', 1)
        if len(parts[0].strip()) <= 6:
            return {'player':parts[1].strip(), 'sport':sport,
                    'color':hex_c, 'traded_from':parts[0].strip()}

    return {'player':val_str, 'sport':sport, 'color':hex_c}

def get_teams_2021(ws):
    teams = []
    for c in range(3, ws.max_column+1):
        team  = cs(ws.cell(2,c).value)
        owner = cs(ws.cell(1,c).value)
        if not team or team.upper() in ('DIR','RND','NFL','NBA','MLB',''): break
        teams.append({'col':c,'owner':owner,'initials':'','team':team})
    return teams

def get_teams_modern(ws):
    STOP = {'DIR','RND','RANK','PLAYER','NFL','NBA','MLB','OWNER','TEAM NAME',
            'INITIALS','INITITAL','INITIAL','NICKNAME','TEAM'}
    teams = []
    for c in range(3, ws.max_column+1):
        r1 = cs(ws.cell(1,c).value); r2 = cs(ws.cell(2,c).value); r3 = cs(ws.cell(3,c).value)
        # Stop at recognized stop words
        if r3.upper() in STOP: break
        # Skip blank separator columns (don't break — 2022 has a blank col before teams)
        if not r3: continue
        owner = r1; initials = r2
        if '(' in r1 and ')' in r1:
            ext = extract_init(r1); owner = clean_owner(r1)
            if ext: initials = ext
        teams.append({'col':c,'owner':owner,'initials':initials,'team':r3})
    return teams

def parse_sheet(ws, year):
    is_2021   = (year == 2021)
    is_future = (year == 2026)
    teams = get_teams_2021(ws) if is_2021 else get_teams_modern(ws)
    picks_start = 3 if is_2021 else 4
    picks, keepers, supplemental, cannot_keep = [], [], [], []
    in_ck = False; keeper_idx = 0

    for row in range(picks_start, ws.max_row+1):
        r_raw = ws.cell(row,1).value; d_raw = cs(ws.cell(row,2).value)
        c3_text = cs(ws.cell(row,3).value).upper()
        round_num = None
        try: round_num = int(r_raw)
        except: pass

        if round_num is None and 'CANNOT' in c3_text and 'KEPT' in c3_text:
            in_ck = True; continue

        if in_ck:
            if round_num is not None:
                in_ck = False
            else:
                if c3_text in {'NFL','NBA','MLB','CANNOT BE KEPT PLAYERS',''}:
                    in_ck = False; continue
                row_vals = []
                for t in teams:
                    cell = ws.cell(row, t['col']); val = cs(cell.value)
                    if val and val.upper() not in {'NFL','NBA','MLB','CANNOT BE KEPT PLAYERS',''}:
                        hex_c = get_hex(cell)
                        row_vals.append({'player':val,'sport':sport_of(hex_c),'color':hex_c})
                    else: row_vals.append(None)
                if len([v for v in row_vals if v]) >= max(1, len(teams)//2):
                    cannot_keep.append(row_vals)
                continue

        if round_num is None: continue

        d_up = d_raw.upper()
        # Valid direction values: arrows, KPR (old), Keeper (new), or blank
        if d_raw and d_up not in ('--->', '<---', 'KPR', 'KEEPER', ''): continue

        explicit_kpr     = (d_up in ('KPR', 'KEEPER'))
        keeper_by_round  = (year in KEEPER_YEARS and round_num in KEEPER_ROUNDS)
        is_kpr  = explicit_kpr or keeper_by_round
        is_supp = (round_num >= 42)

        row_picks = []; any_pick = False
        for t in teams:
            cell = ws.cell(row, t['col']); val = cs(cell.value)
            if not val or val.upper() in ('SUPPLEMENTAL SLOTS', 'KEEPER'):
                row_picks.append(None); continue
            any_pick = True
            p = parse_cell(cell, val, is_future=is_future)
            if is_kpr and not p.get('is_keeper'): p['is_keeper'] = True
            row_picks.append(p)

        row_data = {'round':round_num, 'dir':d_raw, 'picks':row_picks}
        if is_supp:
            supplemental.append(row_data)
        elif is_kpr:
            row_data['keeper_sport'] = keeper_sport(keeper_idx)
            row_data['keeper_num']   = (keeper_idx % 3) + 1
            keepers.append(row_data)
            keeper_idx += 1
        else:
            picks.append(row_data)

    return {'year':year, 'is_future':is_future, 'teams':teams,
            'picks':picks, 'keepers':keepers, 'supplemental':supplemental,
            'cannot_keep':cannot_keep}

def apply_overrides(drafts):
    # 2023 R6 TC = Myles Turner (NBA) — orange cell was a coloring error
    d23 = drafts.get(2023)
    if d23:
        tc_idx = next((i for i, t in enumerate(d23['teams']) if t['initials']=='TC'), None)
        if tc_idx is not None:
            for row in d23['picks']:
                if row['round'] == 6:
                    p = row['picks'][tc_idx]
                    if p and p.get('player') == 'Myles Turner':
                        p['sport'] = 'NBA'; p['color'] = 'C9DAF8'
                        print("  Override: 2023 R6 TC Myles Turner → NBA", file=sys.stderr)

    # 2022 CQ keeper round 33 = "Josh Allen" (Excel cell is truncated to "Josh Alle")
    d22 = drafts.get(2022)
    if d22:
        for row in d22['keepers']:
            for p in row['picks']:
                if p and p.get('player') == 'Josh Alle':
                    p['player'] = 'Josh Allen'
                    print("  Override: 2022 keeper CQ 'Josh Alle' → 'Josh Allen'", file=sys.stderr)

def run():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    drafts = {}
    for sn in wb.sheetnames:
        m = re.search(r'(\d{4})', sn)
        if not m: continue
        year = int(m.group(1))
        d = parse_sheet(wb[sn], year)
        drafts[year] = d
        print(f"  {year}: {len(d['teams'])} teams, {len(d['picks'])} picks, "
              f"{len(d['keepers'])} keepers, {len(d['supplemental'])} supp", file=sys.stderr)
    apply_overrides(drafts)
    return drafts

if __name__ == '__main__':
    d = run()
    print(json.dumps(d, indent=2))
